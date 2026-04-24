import pandas as pd
import io
from datetime import datetime
from flask import send_file, make_response
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import json
class ReportService:
    
    # ==================== EXPORTAR REPORTES ====================
    
    @staticmethod
    def export_pacientes_to_excel(pacientes, nombre_archivo=None):
        """Exporta lista de pacientes a Excel"""
        if not nombre_archivo:
            nombre_archivo = f"pacientes_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        # Crear DataFrame
        data = []
        for p in pacientes:
            data.append({
                'ID': p.idPaciente,
                'Nombre': p.nombre,
                'Apellido Paterno': p.apP,
                'Apellido Materno': p.apM or '',
                'Nombre Completo': p.nombre_completo,
                'Sexo': 'Masculino' if p.sexo == 'M' else 'Femenino' if p.sexo == 'F' else 'Otro',
                'Fecha Nacimiento': p.edadNac.strftime('%d/%m/%Y') if p.edadNac else '',
                'Edad': p.edad if p.edad else '',
                'Teléfono': p.telefono or '',
                'Fecha Registro': datetime.now().strftime('%d/%m/%Y %H:%M')
            })
        
        df = pd.DataFrame(data)
        
        # Crear archivo Excel
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Pacientes', index=False)
            
            # Formatear hoja
            worksheet = writer.sheets['Pacientes']
            ReportService._format_excel_worksheet(worksheet, df)
        
        output.seek(0)
        
        return send_file(
            output,
            as_attachment=True,
            download_name=nombre_archivo,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    
    @staticmethod
    def export_consultas_to_excel(consultas, paciente_nombre, nombre_archivo=None):
        """Exporta consultas de un paciente a Excel"""
        if not nombre_archivo:
            nombre_archivo = f"consultas_{paciente_nombre}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        data = []
        for c in consultas:
            data.append({
                'ID Consulta': c.idConsulta,
                'Fecha': c.fechaConsulta.strftime('%d/%m/%Y') if c.fechaConsulta else '',
                'Hora Entrada': c.horaE.strftime('%H:%M') if c.horaE else '',
                'Hora Salida': c.horaS.strftime('%H:%M') if c.horaS else '',
                'Descripción': c.descripcion or '',
                'Peso (kg)': c.diagnostico.peso if c.diagnostico else '',
                'Estatura (m)': c.diagnostico.estatura if c.diagnostico else '',
                'IMC': c.diagnostico.imc if c.diagnostico else '',
                'Diagnóstico': c.diagnostico.descripcion if c.diagnostico else '',
                'Planes Alimenticios': ReportService._get_planes_text(c)
            })
        
        df = pd.DataFrame(data)
        
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=f'Consultas_{paciente_nombre}', index=False)
            worksheet = writer.sheets[f'Consultas_{paciente_nombre}']
            ReportService._format_excel_worksheet(worksheet, df)
        
        output.seek(0)
        
        return send_file(
            output,
            as_attachment=True,
            download_name=nombre_archivo,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    
    @staticmethod
    def export_resumen_nutriologo_to_excel(pacientes, estadisticas, nombre_archivo=None):
        """Exporta resumen completo del nutriólogo a Excel con múltiples hojas"""
        if not nombre_archivo:
            nombre_archivo = f"reporte_nutriologo_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Hoja 1: Resumen
            resumen_data = [
                {'Métrica': 'Total Pacientes', 'Valor': estadisticas['total_pacientes']},
                {'Métrica': 'Total Consultas', 'Valor': estadisticas['consultas_totales']},
                {'Métrica': 'Edad Promedio', 'Valor': f"{estadisticas['edad_promedio']} años"},
                {'Métrica': 'Edad Mínima', 'Valor': f"{estadisticas['edad_min']} años"},
                {'Métrica': 'Edad Máxima', 'Valor': f"{estadisticas['edad_max']} años"},
                {'Métrica': 'Pacientes Masculinos', 'Valor': estadisticas['generos'].get('M', 0)},
                {'Métrica': 'Pacientes Femeninos', 'Valor': estadisticas['generos'].get('F', 0)},
                {'Métrica': 'Fecha Reporte', 'Valor': datetime.now().strftime('%d/%m/%Y %H:%M')}
            ]
            df_resumen = pd.DataFrame(resumen_data)
            df_resumen.to_excel(writer, sheet_name='Resumen', index=False)
            
            # Hoja 2: Pacientes
            pacientes_data = []
            for p in pacientes:
                pacientes_data.append({
                    'ID': p.idPaciente,
                    'Nombre': p.nombre,
                    'Apellido': p.apP,
                    'Sexo': p.sexo,
                    'Edad': p.edad,
                    'Teléfono': p.telefono or ''
                })
            df_pacientes = pd.DataFrame(pacientes_data)
            df_pacientes.to_excel(writer, sheet_name='Pacientes', index=False)
            
            # Hoja 3: Distribución por Edad
            edad_data = [{'Rango': k, 'Cantidad': v} for k, v in estadisticas['rangos_edad'].items()]
            df_edad = pd.DataFrame(edad_data)
            df_edad.to_excel(writer, sheet_name='Distribución Edad', index=False)
            
            # Formatear todas las hojas
            for sheet_name in writer.sheets:
                worksheet = writer.sheets[sheet_name]
                if sheet_name == 'Resumen':
                    ReportService._format_resumen_worksheet(worksheet)
                else:
                    df = pd.read_excel(output, sheet_name=sheet_name) if output else None
                    ReportService._format_excel_worksheet(worksheet, df)
        
        output.seek(0)
        
        return send_file(
            output,
            as_attachment=True,
            download_name=nombre_archivo,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    
    # ==================== IMPORTAR REPORTES ====================
    
    @staticmethod
    def import_pacientes_from_excel(file, nutriologo_id):
        """Importa pacientes desde archivo Excel"""
        try:
            df = pd.read_excel(file)
            
            # Validar columnas requeridas
            required_columns = ['Nombre', 'Apellido Paterno', 'Sexo', 'Fecha Nacimiento']
            for col in required_columns:
                if col not in df.columns:
                    return {'error': f'Falta columna requerida: {col}'}
            
            resultados = {
                'exitosos': [],
                'errores': [],
                'total': len(df),
                'importados': 0
            }
            
            for idx, row in df.iterrows():
                try:
                    # Procesar fecha
                    fecha_nac = row['Fecha Nacimiento']
                    if isinstance(fecha_nac, str):
                        fecha_nac = datetime.strptime(fecha_nac, '%d/%m/%Y').date()
                    
                    # Crear paciente
                    paciente_id = DatabaseService.create_paciente(
                        nombre=row['Nombre'],
                        apP=row['Apellido Paterno'],
                        apM=row.get('Apellido Materno', ''),
                        sexo=row['Sexo'][0] if len(row['Sexo']) > 0 else 'M',
                        edadNac=fecha_nac,
                        telefono=row.get('Teléfono', ''),
                        nutriologo_id=nutriologo_id
                    )
                    
                    resultados['exitosos'].append({
                        'fila': idx + 2,
                        'nombre': row['Nombre'],
                        'id': paciente_id
                    })
                    resultados['importados'] += 1
                    
                except Exception as e:
                    resultados['errores'].append({
                        'fila': idx + 2,
                        'nombre': row.get('Nombre', 'Desconocido'),
                        'error': str(e)
                    })
            
            return resultados
            
        except Exception as e:
            return {'error': f'Error al leer archivo: {str(e)}'}
    
    
    @staticmethod
    def import_consultas_from_excel(file, paciente_id, nutriologo_id):
      """Importa consultas desde archivo Excel"""
      try:
         df = pd.read_excel(file)
         
         required_columns = ['Fecha', 'Hora Entrada', 'Hora Salida', 'Peso (kg)', 'Estatura (m)']
         for col in required_columns:
               if col not in df.columns:
                  return {'error': f'Falta columna requerida: {col}'}
         
         resultados = {
               'exitosos': [],
               'errores': [],
               'total': len(df),
               'importados': 0
         }
         
         for idx, row in df.iterrows():
               try:
                  # Procesar fecha
                  fecha = row['Fecha']
                  if isinstance(fecha, str):
                     fecha = datetime.strptime(fecha, '%d/%m/%Y').date()
                  elif hasattr(fecha, 'date'):
                     fecha = fecha.date()
                  
                  # Procesar horas
                  hora_e = row['Hora Entrada']
                  if isinstance(hora_e, str):
                     hora_e = datetime.strptime(hora_e, '%H:%M').time()
                  elif hasattr(hora_e, 'time'):
                     hora_e = hora_e.time()
                  
                  hora_s = row['Hora Salida']
                  if isinstance(hora_s, str):
                     hora_s = datetime.strptime(hora_s, '%H:%M').time()
                  elif hasattr(hora_s, 'time'):
                     hora_s = hora_s.time()
                  
                  # Crear consulta
                  consulta_id = DatabaseService.create_consulta(
                     fechaConsulta=fecha,
                     horaE=hora_e,
                     horaS=hora_s,
                     descripcion=row.get('Descripción', ''),
                     paciente_id=paciente_id
                  )
                  
                  # Calcular IMC
                  peso = float(row['Peso (kg)'])
                  estatura = float(row['Estatura (m)'])
                  imc = round(peso / (estatura ** 2), 2)
                  
                  # Crear diagnóstico
                  DatabaseService.create_diagnostico(
                     peso=peso,
                     estatura=estatura,
                     imc=imc,
                     descripcion=row.get('Diagnóstico', ''),
                     consulta_id=consulta_id,
                     nutriologo_id=nutriologo_id
                  )
                  
                  resultados['exitosos'].append({
                     'fila': idx + 2,
                     'fecha': fecha.strftime('%d/%m/%Y'),
                     'id': consulta_id
                  })
                  resultados['importados'] += 1
                  
               except Exception as e:
                  resultados['errores'].append({
                     'fila': idx + 2,
                     'error': str(e)
                  })
        
         return resultados
         
      except Exception as e:
         return {'error': f'Error al leer archivo: {str(e)}'}
    
    # ==================== MÉTODOS AUXILIARES ====================
    
    @staticmethod
    def _get_planes_text(consulta):
        """Obtiene texto de planes alimenticios"""
        if hasattr(consulta, 'planes') and consulta.planes:
            planes_text = []
            for p in consulta.planes:
                planes_text.append(f"{p.kcalD} kcal ({p.fechaI} - {p.fechaF})")
            return '; '.join(planes_text)
        return ''
    
    @staticmethod
    def _format_excel_worksheet(worksheet, df):
        """Formatea hoja de Excel con estilos profesionales"""
        # Estilos
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Formatear encabezados
        for col in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Ajustar ancho de columnas
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    @staticmethod
    def _format_resumen_worksheet(worksheet):
        """Formatea hoja de resumen"""
        # Estilo para métricas
        metric_font = Font(bold=True, size=12)
        value_font = Font(size=12)
        
        for row in range(1, worksheet.max_row + 1):
            cell_metric = worksheet.cell(row=row, column=1)
            cell_value = worksheet.cell(row=row, column=2)
            
            cell_metric.font = metric_font
            cell_value.font = value_font
            
            # Color para totales
            if 'Total' in str(cell_metric.value):
                cell_value.font = Font(bold=True, color="2E75B6", size=14)
        
        # Ajustar ancho
        worksheet.column_dimensions['A'].width = 25
        worksheet.column_dimensions['B'].width = 20